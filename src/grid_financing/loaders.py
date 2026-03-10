from __future__ import annotations

import calendar
import logging
import re
import unicodedata
from pathlib import Path
from typing import Any, Iterable

logger = logging.getLogger(__name__)

import pandas as pd
from openpyxl import load_workbook

from .source_registry import (
    PROJECT_ROOT,
    SOURCE_REGISTRY,
    WORKBOOK_SIGNATURES,
    ensure_columns_present,
    resolve_source,
)

STATUS_MAP = {
    1: "Under Consideration",
    2: "Planning",
    3: "Permitting",
    4: "Construction",
    5: "Completed or Other",
}

COUNTRY_CODE_TO_ISO3 = {
    "AL": "ALB",
    "AT": "AUT",
    "BA": "BIH",
    "BE": "BEL",
    "BG": "BGR",
    "CH": "CHE",
    "CY": "CYP",
    "CZ": "CZE",
    "DE": "DEU",
    "DK": "DNK",
    "EE": "EST",
    "ES": "ESP",
    "FI": "FIN",
    "FR": "FRA",
    "GB": "GBR",
    "GR": "GRC",
    "HR": "HRV",
    "HU": "HUN",
    "IE": "IRL",
    "IT": "ITA",
    "LT": "LTU",
    "LU": "LUX",
    "LV": "LVA",
    "ME": "MNE",
    "MK": "MKD",
    "NL": "NLD",
    "NO": "NOR",
    "PL": "POL",
    "PT": "PRT",
    "RO": "ROU",
    "RS": "SRB",
    "SE": "SWE",
    "SI": "SVN",
    "SK": "SVK",
    "TR": "TUR",
    "UA": "UKR",
}


def normalize_header_fragment(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.lower().startswith("unnamed:"):
        return ""
    text = text.replace("Δ", "d")
    text = text.replace("…", "")
    text = text.replace("%", " pct ")
    text = text.replace("&", " and ")
    text = text.replace("/", " ")
    text = text.replace("-", " ")
    text = text.replace("\n", " ")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^0-9A-Za-z_ ]+", " ", text)
    text = re.sub(r"\s+", "_", text)
    return text.strip("_").lower()


def normalize_headers(header_rows: list[tuple[Any, ...]]) -> list[str]:
    width = max(len(row) for row in header_rows)
    columns: list[str] = []
    seen: dict[str, int] = {}
    for column_index in range(width):
        parts = [
            normalize_header_fragment(row[column_index] if column_index < len(row) else None)
            for row in header_rows
        ]
        parts = [part for part in parts if part]
        column_name = "_".join(parts) if parts else f"column_{column_index + 1}"
        count = seen.get(column_name, 0)
        seen[column_name] = count + 1
        columns.append(column_name if count == 0 else f"{column_name}_{count + 1}")
    return columns


def read_excel_table(
    path: Path,
    sheet_name: str,
    *,
    header_rows: tuple[int, ...],
    data_start_row: int,
) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"{path.name} is missing required sheet `{sheet_name}`")
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    selected_headers = [rows[row_index - 1] for row_index in header_rows]
    columns = normalize_headers(selected_headers)
    frame = pd.DataFrame(rows[data_start_row - 1 :], columns=columns)
    frame = frame.dropna(axis=1, how="all").dropna(axis=0, how="all").reset_index(drop=True)
    return frame


def _coerce_numeric_columns(df: pd.DataFrame, numeric_columns: Iterable[str]) -> pd.DataFrame:
    for column in numeric_columns:
        if column in df.columns:
            df[column] = df[column].apply(_parse_numeric_value)
    return df


def _parse_numeric_value(value: Any) -> float | pd.NA:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return pd.NA
    if isinstance(value, (int, float)):
        return float(value)
    matches = re.findall(r"-?\d+(?:\.\d+)?", str(value).replace(",", " "))
    if not matches:
        return pd.NA
    numbers = [float(match) for match in matches]
    if len(numbers) == 1 or all(number == numbers[0] for number in numbers):
        return numbers[0]
    logger.debug(
        "Multi-value numeric cell %r resolved to max(%s) = %s",
        value,
        numbers,
        max(numbers),
    )
    return max(numbers)


def _resolve_source_path(source_id: str) -> tuple[Path, str]:
    resolved = resolve_source(source_id)
    if not resolved.exists or resolved.path is None:
        raise FileNotFoundError(resolved.descriptor.missing_message)
    return resolved.path, resolved.variant


def _validate_sheet_signature(df: pd.DataFrame, workbook_key: str, sheet_name: str) -> None:
    required_headers = WORKBOOK_SIGNATURES[workbook_key][sheet_name]
    ensure_columns_present(df.columns, required_headers, f"{workbook_key}:{sheet_name}")


def load_transmission_projects() -> pd.DataFrame:
    path, variant = _resolve_source_path("tyndp2024_workbook")
    df = read_excel_table(path, "Trans.Projects", header_rows=(2,), data_start_row=3)
    _validate_sheet_signature(df, "tyndp2024", "Trans.Projects")
    df = _coerce_numeric_columns(
        df,
        (
            "project_id",
            "status_id_1_under_consideration_2_in_planning_but_not_permitting_3_in_permitting_4_under_construction",
            "transfer_capacity_increase_a_b_mw",
            "transfer_capacity_increase_b_a_mw",
        ),
    )
    df["source_variant_tyndp2024"] = variant
    df["source_path_tyndp2024"] = str(path)
    df["is_cross_border"] = df["is_the_project_cross_border"].astype(str).str.lower().eq("true")
    df["is_internal"] = df.get("is_the_project_internal", pd.Series(index=df.index, dtype=object)).astype(str).str.lower().eq("true")
    df["status"] = (
        df["status_id_1_under_consideration_2_in_planning_but_not_permitting_3_in_permitting_4_under_construction"]
        .map(STATUS_MAP)
        .fillna("Missing")
    )
    df["countries_list"] = df["country"].apply(parse_country_codes)
    df["countries"] = df["countries_list"].apply(lambda values: " ; ".join(values))
    df["country_count"] = df["countries_list"].apply(len)
    df["is_multi_country"] = df["country_count"] > 2
    df["country_a"] = df["countries_list"].apply(lambda values: values[0] if len(values) >= 1 else pd.NA)
    df["country_b"] = df["countries_list"].apply(lambda values: values[1] if len(values) >= 2 else pd.NA)
    df.loc[df["country_count"] != 2, ["country_a", "country_b"]] = pd.NA
    df["capacity_mw"] = df[["transfer_capacity_increase_a_b_mw", "transfer_capacity_increase_b_a_mw"]].max(axis=1, skipna=True)
    return df


def load_transmission_investments() -> pd.DataFrame:
    path, variant = _resolve_source_path("tyndp2024_workbook")
    df = read_excel_table(path, "Trans.Investments", header_rows=(2,), data_start_row=3)
    _validate_sheet_signature(df, "tyndp2024", "Trans.Investments")
    df = _coerce_numeric_columns(
        df,
        (
            "investment_number",
            "this_investment_belongs_to_project_number",
            "capacity_of_the_investment_mw",
            "estimated_capex_meur",
        ),
    )
    df["source_variant_tyndp2024"] = variant
    return df


def load_tyndp_2024_cba() -> pd.DataFrame:
    path, variant = _resolve_source_path("tyndp2024_workbook")
    outputs = []
    for sheet_name, prefix in (
        ("2030NT-EU27", "2030nt_eu27"),
        ("2040NT-EU27", "2040nt_eu27"),
        ("2040DE-EU27", "2040de_eu27"),
    ):
        df = read_excel_table(path, sheet_name, header_rows=(1, 2), data_start_row=4)
        _validate_sheet_signature(df, "tyndp2024", sheet_name)
        df = _coerce_numeric_columns(
            df,
            (
                "project_id",
                "dsew_weighted_avg",
                "dsew_co2_weighted_avg",
                "dres_weighted_avg",
                "dco2_market_weighted_avg",
            ),
        )
        df = _collapse_project_rows(df)
        outputs.append(
            df[
                [
                    "project_id",
                    "project_name",
                    "dsew_weighted_avg",
                    "dsew_co2_weighted_avg",
                    "dres_weighted_avg",
                    "dco2_market_weighted_avg",
                ]
            ].rename(
                columns={
                    "dsew_weighted_avg": f"dsew_{prefix}_meur_per_year",
                    "dsew_co2_weighted_avg": f"dsew_co2_{prefix}_meur_per_year",
                    "dres_weighted_avg": f"dres_{prefix}_gwh_per_year",
                    "dco2_market_weighted_avg": f"dco2_market_{prefix}_ktonnes_per_year",
                }
            )
        )

    merged = outputs[0]
    for item in outputs[1:]:
        merged = merged.merge(item.drop(columns=["project_name"]), on="project_id", how="outer")
    merged["source_variant_tyndp2024"] = variant
    return merged


def load_tyndp_2022_cba() -> pd.DataFrame:
    path, variant = _resolve_source_path("tyndp2022_workbook")
    outputs = []
    for sheet_name, prefix in (
        ("NT2030_EU", "nt2030_2022"),
        ("DE2030_EU", "de2030_2022"),
        ("DE2040_EU", "de2040_2022"),
    ):
        df = read_excel_table(path, sheet_name, header_rows=(3, 4), data_start_row=6)
        _validate_sheet_signature(df, "tyndp2022", sheet_name)
        df = _coerce_numeric_columns(
            df,
            (
                "project_id",
                "dsew_weighted_avg",
                "dsew_co2_weighted_avg",
                "dres_weighted_avg",
                "dco2_market_weighted_avg",
            ),
        )
        df = _collapse_project_rows(df)
        outputs.append(
            df[
                [
                    "project_id",
                    "project_name",
                    "dsew_weighted_avg",
                    "dsew_co2_weighted_avg",
                    "dres_weighted_avg",
                    "dco2_market_weighted_avg",
                ]
            ].rename(
                columns={
                    "dsew_weighted_avg": f"dsew_{prefix}_meur_per_year",
                    "dsew_co2_weighted_avg": f"dsew_co2_{prefix}_meur_per_year",
                    "dres_weighted_avg": f"dres_{prefix}_gwh_per_year",
                    "dco2_market_weighted_avg": f"dco2_market_{prefix}_ktonnes_per_year",
                }
            )
        )

    merged = outputs[0]
    for item in outputs[1:]:
        merged = merged.merge(item.drop(columns=["project_name"]), on="project_id", how="outer")
    merged["source_variant_tyndp2022"] = variant
    return merged


def aggregate_project_capex(investments_df: pd.DataFrame) -> pd.DataFrame:
    grouped = (
        investments_df.groupby("this_investment_belongs_to_project_number", dropna=False)
        .agg(
            capex_meur=("estimated_capex_meur", "sum"),
            investment_count=("investment_number", "count"),
        )
        .reset_index()
        .rename(columns={"this_investment_belongs_to_project_number": "project_id"})
    )
    return grouped


def parse_country_codes(value: Any) -> list[str]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return []
    return [item.strip() for item in re.split(r"[;/,]", str(value)) if item and item.strip()]


def _join_unique_strings(values: pd.Series) -> str:
    return " | ".join(
        dict.fromkeys(str(value).strip() for value in values if pd.notna(value) and str(value).strip())
    )


def _first_non_null(values: pd.Series) -> Any:
    return next(
        (value for value in values if pd.notna(value) and str(value).strip()),
        pd.NA,
    )


def _collapse_project_rows(df: pd.DataFrame) -> pd.DataFrame:
    if "project_id" not in df.columns:
        return df
    df = df[df["project_id"].notna()].copy()
    if df.empty:
        return df

    aggregations: dict[str, Any] = {}
    for column in df.columns:
        if column == "project_id":
            continue
        if pd.api.types.is_numeric_dtype(df[column]):
            aggregations[column] = "max"
        elif column == "project_name":
            aggregations[column] = _join_unique_strings
        else:
            aggregations[column] = _first_non_null

    return df.groupby("project_id", as_index=False).agg(aggregations)


def load_manual_csv(source_id: str) -> pd.DataFrame:
    resolved = resolve_source(source_id)
    schema = SOURCE_REGISTRY[source_id].schema
    if not resolved.exists or resolved.path is None:
        return pd.DataFrame(columns=list(schema))
    df = pd.read_csv(resolved.path)
    ensure_columns_present(df.columns, schema, source_id)
    return df


def append_flag(df: pd.DataFrame, mask: pd.Series, flag: str) -> pd.DataFrame:
    existing = df["data_quality_flags"].fillna("")
    df.loc[mask, "data_quality_flags"] = existing[mask].apply(lambda value: _merge_flag_string(value, flag))
    return df


def _merge_flag_string(existing: str, flag: str) -> str:
    flags = {item for item in existing.split(";") if item}
    flags.add(flag)
    return ";".join(sorted(flags))


def _apply_project_overrides(project_df: pd.DataFrame, overrides_df: pd.DataFrame) -> pd.DataFrame:
    if overrides_df.empty:
        return project_df
    overrides = overrides_df.copy()
    overrides["project_id"] = pd.to_numeric(overrides["project_id"], errors="coerce")
    merged = project_df.merge(overrides, on="project_id", how="left")

    merged.loc[merged["countries_override"].notna(), "countries_list"] = merged.loc[
        merged["countries_override"].notna(), "countries_override"
    ].apply(parse_country_codes)
    merged["countries"] = merged["countries_list"].apply(lambda values: " ; ".join(values))
    merged["country_count"] = merged["countries_list"].apply(len)
    merged["is_multi_country"] = merged["country_count"] > 2

    merged.loc[merged["primary_border_a"].notna(), "country_a"] = merged["primary_border_a"]
    merged.loc[merged["primary_border_b"].notna(), "country_b"] = merged["primary_border_b"]
    merged.loc[merged["transfer_capacity_ab_mw_override"].notna(), "transfer_capacity_increase_a_b_mw"] = pd.to_numeric(
        merged["transfer_capacity_ab_mw_override"], errors="coerce"
    )
    merged.loc[merged["transfer_capacity_ba_mw_override"].notna(), "transfer_capacity_increase_b_a_mw"] = pd.to_numeric(
        merged["transfer_capacity_ba_mw_override"], errors="coerce"
    )
    merged["capacity_mw"] = merged[["transfer_capacity_increase_a_b_mw", "transfer_capacity_increase_b_a_mw"]].max(axis=1, skipna=True)
    return merged


def _apply_participant_data(project_df: pd.DataFrame, participants_df: pd.DataFrame) -> pd.DataFrame:
    if participants_df.empty:
        project_df["participant_count"] = 0
        project_df["capex_share_method"] = project_df["country_count"].apply(lambda count: "default-50-50" if count == 2 else "missing")
        return project_df

    participants = participants_df.copy()
    participants["project_id"] = pd.to_numeric(participants["project_id"], errors="coerce")
    participants["capex_share_pct"] = pd.to_numeric(participants["capex_share_pct"], errors="coerce")
    participants["participant_order"] = pd.to_numeric(participants["participant_order"], errors="coerce")
    counts = participants.groupby("project_id").size().rename("participant_count")
    merged = project_df.merge(counts, on="project_id", how="left")
    merged["participant_count"] = merged["participant_count"].fillna(0).astype(int)
    merged["capex_share_method"] = merged["participant_count"].apply(lambda count: "participant-table" if count else "default-50-50")
    return merged


def _attach_credit_reference(project_df: pd.DataFrame, participants_df: pd.DataFrame, credit_df: pd.DataFrame) -> pd.DataFrame:
    df = project_df.copy()
    if credit_df.empty:
        for column in (
            "tso_a_name",
            "tso_b_name",
            "tso_a_rating",
            "tso_b_rating",
            "tso_a_rab_beur",
            "tso_b_rab_beur",
            "sovereign_a_rating",
            "sovereign_b_rating",
            "sovereign_a_debt_to_gdp_pct",
            "sovereign_b_debt_to_gdp_pct",
            "sovereign_a_deficit_to_gdp_pct",
            "sovereign_b_deficit_to_gdp_pct",
            "cohesion_country_a",
            "cohesion_country_b",
        ):
            df[column] = pd.NA
        return df

    credit = credit_df.copy()
    credit["country_code"] = credit["country_code"].astype(str).str.strip().str.upper()
    credit["tso_rab_beur"] = pd.to_numeric(credit["tso_rab_beur"], errors="coerce")
    credit["debt_to_gdp_pct"] = pd.to_numeric(credit["debt_to_gdp_pct"], errors="coerce")
    credit["deficit_to_gdp_pct"] = pd.to_numeric(credit["deficit_to_gdp_pct"], errors="coerce")
    credit["cohesion_country"] = credit["cohesion_country"].astype(str).str.lower().isin({"true", "1", "yes", "y"})

    if not participants_df.empty:
        participants = participants_df.copy()
        participants["project_id"] = pd.to_numeric(participants["project_id"], errors="coerce")
        participants["country_code"] = participants["country_code"].astype(str).str.strip().str.upper()
        participants["capex_share_pct"] = pd.to_numeric(participants["capex_share_pct"], errors="coerce")
        participants["participant_order"] = pd.to_numeric(participants["participant_order"], errors="coerce")
        participants = participants.sort_values(["project_id", "participant_order", "country_code"])
        side_participants = []
        for side_index, side in enumerate(("a", "b"), start=1):
            item = participants.groupby("project_id").nth(side_index - 1).reset_index()
            item = item.rename(
                columns={
                    "country_code": f"country_{side}",
                    "tso_name": f"tso_{side}_name",
                    "capex_share_pct": f"capex_share_pct_{side}",
                }
            )
            side_participants.append(item[["project_id", f"country_{side}", f"tso_{side}_name", f"capex_share_pct_{side}"]])
        for item in side_participants:
            df = df.merge(item, on="project_id", how="left", suffixes=("", "_participant"))

    for side in ("a", "b"):
        country_column = f"country_{side}"
        if country_column not in df.columns:
            df[country_column] = pd.NA
        side_credit = credit.drop_duplicates("country_code").rename(
            columns={
                "country_code": country_column,
                "tso_name": f"tso_{side}_name_reference",
                "tso_credit_rating": f"tso_{side}_rating",
                "tso_rab_beur": f"tso_{side}_rab_beur",
                "sovereign_rating": f"sovereign_{side}_rating",
                "debt_to_gdp_pct": f"sovereign_{side}_debt_to_gdp_pct",
                "deficit_to_gdp_pct": f"sovereign_{side}_deficit_to_gdp_pct",
                "cohesion_country": f"cohesion_country_{side}",
            }
        )
        df = df.merge(
            side_credit[
                [
                    country_column,
                    f"tso_{side}_name_reference",
                    f"tso_{side}_rating",
                    f"tso_{side}_rab_beur",
                    f"sovereign_{side}_rating",
                    f"sovereign_{side}_debt_to_gdp_pct",
                    f"sovereign_{side}_deficit_to_gdp_pct",
                    f"cohesion_country_{side}",
                ]
            ],
            on=country_column,
            how="left",
        )
        if f"tso_{side}_name" not in df.columns:
            df[f"tso_{side}_name"] = df[f"tso_{side}_name_reference"]

    df["capex_share_pct_a"] = pd.to_numeric(df.get("capex_share_pct_a"), errors="coerce")
    df["capex_share_pct_b"] = pd.to_numeric(df.get("capex_share_pct_b"), errors="coerce")
    two_country_mask = df["country_count"].eq(2)
    df.loc[two_country_mask & df["capex_share_pct_a"].isna(), "capex_share_pct_a"] = 50.0
    df.loc[two_country_mask & df["capex_share_pct_b"].isna(), "capex_share_pct_b"] = 50.0
    df["project_capex_share_a_meur"] = df["capex_meur"] * df["capex_share_pct_a"] / 100
    df["project_capex_share_b_meur"] = df["capex_meur"] * df["capex_share_pct_b"] / 100
    return df


def load_local_hourly_price_data() -> pd.DataFrame:
    path, variant = _resolve_source_path("local_hourly_prices")
    df = pd.read_csv(path)
    ensure_columns_present(df.columns, SOURCE_REGISTRY["local_hourly_prices"].schema, "local_hourly_prices")
    df["price_eur_per_mwh"] = pd.to_numeric(df["Price (EUR/MWhe)"], errors="coerce")
    df["datetime_utc"] = pd.to_datetime(df["Datetime (UTC)"], errors="coerce", utc=True)
    df["data_year"] = df["datetime_utc"].dt.year
    df["iso3_code"] = df["ISO3 Code"].astype(str).str.upper()
    df["source_variant_local_hourly_prices"] = variant
    return df


def _latest_full_year(series: pd.Series) -> int | None:
    counts = series.dropna().value_counts()
    eligible = sorted(int(year) for year, count in counts.items() if count >= 8_000)
    return eligible[-1] if eligible else None


def _full_year_hour_count(data_year: int) -> int:
    return 8784 if calendar.isleap(data_year) else 8760


def _build_pair_metrics_for_year(price_df: pd.DataFrame, country_a: str, country_b: str, data_year: int) -> dict[str, Any]:
    iso3_a = COUNTRY_CODE_TO_ISO3.get(country_a)
    iso3_b = COUNTRY_CODE_TO_ISO3.get(country_b)
    if not iso3_a or not iso3_b:
        return {"notes": f"Missing ISO3 mapping for {country_a}/{country_b}"}

    pair = price_df[price_df["iso3_code"].isin({iso3_a, iso3_b})].copy()
    if pair.empty:
        return {"notes": f"No local hourly prices found for {country_a}/{country_b}"}

    pair = pair[pair["data_year"] == data_year]
    left = pair[pair["iso3_code"] == iso3_a][["datetime_utc", "price_eur_per_mwh"]].rename(columns={"price_eur_per_mwh": "price_a"})
    right = pair[pair["iso3_code"] == iso3_b][["datetime_utc", "price_eur_per_mwh"]].rename(columns={"price_eur_per_mwh": "price_b"})
    merged = left.merge(right, on="datetime_utc", how="inner").dropna()
    expected_hours = _full_year_hour_count(data_year)
    if len(merged) < expected_hours:
        return {
            "notes": (
                f"No full-year overlapping hourly prices for {country_a}/{country_b} in {data_year} "
                f"({len(merged)}/{expected_hours} hours)"
            )
        }

    abs_diff = (merged["price_a"] - merged["price_b"]).abs()

    return {
        "avg_price_diff_eur_per_mwh": abs_diff.mean(),
        "hourly_abs_price_diff_sum_eur_per_mwh": abs_diff.sum(),
        "hourly_observation_count": int(len(merged)),
        "price_volatility_a_eur_per_mwh": merged["price_a"].std(),
        "price_volatility_b_eur_per_mwh": merged["price_b"].std(),
        "price_correlation": merged["price_a"].corr(merged["price_b"]),
        "directional_flow_a_to_b_share": (merged["price_a"] > merged["price_b"]).mean(),
        "data_year": int(data_year),
        "source_name": "local_hourly_country_proxy",
        "source_type": "local",
        "source_url": str(resolve_source("local_hourly_prices").path),
        "notes": (
            "Development-mode country-level proxy prices from the local hourly dataset "
            "`data/european_wholesale_electricity_price_data_hourly/all_countries.csv`."
        ),
    }


def _build_pair_metrics(price_df: pd.DataFrame, country_a: str, country_b: str) -> dict[str, Any]:
    iso3_a = COUNTRY_CODE_TO_ISO3.get(country_a)
    iso3_b = COUNTRY_CODE_TO_ISO3.get(country_b)
    if not iso3_a or not iso3_b:
        return {"notes": f"Missing ISO3 mapping for {country_a}/{country_b}"}

    pair = price_df[price_df["iso3_code"].isin({iso3_a, iso3_b})].copy()
    if pair.empty:
        return {"notes": f"No local hourly prices found for {country_a}/{country_b}"}

    candidate_years = sorted(int(year) for year in pair["data_year"].dropna().unique())
    for data_year in reversed(candidate_years):
        metrics = _build_pair_metrics_for_year(price_df, country_a, country_b, data_year)
        if pd.notna(metrics.get("avg_price_diff_eur_per_mwh")):
            return metrics
    fallback_year = _latest_full_year(pair["data_year"])
    if fallback_year is not None:
        return _build_pair_metrics_for_year(price_df, country_a, country_b, fallback_year)
    return {"notes": f"No full-year local hourly prices found for {country_a}/{country_b}"}


def build_price_metrics(
    project_df: pd.DataFrame,
    *,
    development_mode: bool = True,
    price_years: Iterable[int] | None = None,
) -> pd.DataFrame:
    manual_price_df = load_manual_csv("dayahead_price_inputs")
    border_map_df = load_manual_csv("border_zone_map")
    if not development_mode and not manual_price_df.empty:
        return manual_price_df.copy()

    if not development_mode and manual_price_df.empty:
        raise FileNotFoundError(SOURCE_REGISTRY["dayahead_price_inputs"].missing_message)

    price_df = load_local_hourly_price_data()
    requested_years = tuple(sorted({int(year) for year in price_years})) if price_years is not None else None
    rows = []
    border_map_lookup = {}
    if not border_map_df.empty:
        border_map = border_map_df.copy()
        border_map["project_id"] = pd.to_numeric(border_map["project_id"], errors="coerce")
        border_map_lookup = border_map.set_index("project_id").to_dict("index")

    for _, row in project_df.iterrows():
        mapped = border_map_lookup.get(row["project_id"], {})
        country_a = mapped.get("country_a") or row.get("country_a")
        country_b = mapped.get("country_b") or row.get("country_b")
        metrics_rows: list[dict[str, Any]]
        if pd.isna(country_a) or pd.isna(country_b) or not str(country_a).strip() or not str(country_b).strip():
            if requested_years is None:
                metrics_rows = [{"notes": "Missing border mapping for price metrics."}]
            else:
                metrics_rows = [
                    {
                        "data_year": data_year,
                        "price_scenario": f"proxy_{data_year}",
                        "notes": "Missing border mapping for price metrics.",
                    }
                    for data_year in requested_years
                ]
        else:
            if requested_years is None:
                metrics = _build_pair_metrics(price_df, str(country_a), str(country_b))
                scenario_name = (
                    f"proxy_{int(metrics['data_year'])}" if pd.notna(metrics.get("data_year")) else "historical_proxy"
                )
                metrics_rows = [{**metrics, "price_scenario": scenario_name}]
            else:
                metrics_rows = []
                for data_year in requested_years:
                    metrics = _build_pair_metrics_for_year(price_df, str(country_a), str(country_b), data_year)
                    metrics_rows.append(
                        {
                            **metrics,
                            "data_year": data_year,
                            "price_scenario": f"proxy_{data_year}",
                        }
                    )
        for metrics in metrics_rows:
            rows.append(
                {
                    "project_id": row["project_id"],
                    "zone_a": mapped.get("zone_a"),
                    "zone_b": mapped.get("zone_b"),
                    "country_a": country_a,
                    "country_b": country_b,
                    **metrics,
                }
            )
    return pd.DataFrame(rows)


def build_project_master_table(*, development_mode: bool = True, price_years: Iterable[int] | None = None) -> pd.DataFrame:
    projects = load_transmission_projects()
    investments = load_transmission_investments()
    cba_2024 = load_tyndp_2024_cba()
    cba_2022 = load_tyndp_2022_cba()
    overrides = load_manual_csv("project_overrides")
    participants = load_manual_csv("project_participants")
    credit_reference = load_manual_csv("tso_credit_reference")

    master = projects[projects["is_cross_border"]].copy()
    master = master.merge(aggregate_project_capex(investments), on="project_id", how="left")
    master = master.merge(cba_2024, on="project_id", how="left")
    master = master.merge(cba_2022, on="project_id", how="left")
    master = _apply_project_overrides(master, overrides)
    master = _apply_participant_data(master, participants)
    master = _attach_credit_reference(master, participants, credit_reference)
    price_metrics = build_price_metrics(master, development_mode=development_mode, price_years=price_years)
    master = master.merge(price_metrics, on="project_id", how="left", suffixes=("", "_price"))

    master["data_quality_flags"] = ""
    master["assumptions_note"] = (
        "Capacity uses the larger directional transfer-capacity value when both are present. "
        "Price metrics default to the local hourly CSV in development mode. "
        "Estimated congestion rent is a majorant based on the sum of absolute hourly price spreads."
    )
    master["price_input_mode"] = "development-local-proxy" if development_mode else "manual-or-external"
    if development_mode and price_years is not None:
        master["price_input_mode"] = "development-local-proxy-yearly-scenarios"

    master = append_flag(
        master,
        master[["transfer_capacity_increase_a_b_mw", "transfer_capacity_increase_b_a_mw"]].isna().all(axis=1),
        "missing_transfer_capacity",
    )
    master = append_flag(master, master["capex_meur"].isna(), "missing_capex")
    master = append_flag(master, master["country_count"].eq(0), "missing_country_mapping")
    master = append_flag(master, master["country_count"].gt(2) & master["participant_count"].eq(0), "multi_country_needs_participants")
    master = append_flag(master, master["price_scenario"].isna() | master["avg_price_diff_eur_per_mwh"].isna(), "missing_price_inputs")
    master = append_flag(master, master["tso_a_rab_beur"].isna() | master["tso_b_rab_beur"].isna(), "missing_credit_reference")
    master["data_quality_flags"] = master["data_quality_flags"].replace("", pd.NA)
    master["has_data_quality_issue"] = master["data_quality_flags"].notna()
    master["analysis_set"] = "cross-border-transmission"
    sort_columns = [column for column in ("project_id", "data_year") if column in master.columns]
    return master.sort_values(sort_columns).reset_index(drop=True)


def pipeline_report(project_df: pd.DataFrame) -> dict[str, Any]:
    return {
        "cross_border_count": int(project_df["project_id"].nunique()),
        "blank_transfer_capacity_count": int(project_df["data_quality_flags"].fillna("").str.contains("missing_transfer_capacity").sum()),
        "unmapped_multi_country_count": int(project_df["data_quality_flags"].fillna("").str.contains("multi_country_needs_participants").sum()),
        "workbook_variant_2024": project_df["source_variant_tyndp2024"].dropna().iloc[0] if not project_df.empty else "",
    }
